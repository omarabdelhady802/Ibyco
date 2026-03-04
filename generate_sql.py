"""
Reads data/vehicles.xlsx and generates Ibyco/schema.sql
with CREATE TABLE statements + INSERT statements.

Usage:
    python Ibyco/generate_sql.py
"""

import openpyxl
import os

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(BASE_DIR, "data", "vehicles.xlsx")
SQL_PATH = os.path.join(BASE_DIR, "Ibyco", "schema.sql")

SCHEMA = """\
-- Ibyco Database Schema
-- Auto-generated from vehicles.xlsx

-- =========================
-- Users Table
-- =========================
CREATE TABLE IF NOT EXISTS users (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name VARCHAR(50) NOT NULL UNIQUE,
    password VARCHAR(50) NOT NULL
);

-- =========================
-- Clients Table
-- =========================
CREATE TABLE IF NOT EXISTS clients (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    phone_number VARCHAR(30) NOT NULL UNIQUE,
    chat_summary TEXT,
    last_user_reply TEXT,
    last_bot_reply TEXT,
    last_bot_reply_type VARCHAR(50),
    last_user_message_at DATETIME,
    last_bot_message_at DATETIME,
    info TEXT,
    has_purchased BOOLEAN DEFAULT 0,
    purchase_date DATETIME,
    is_active BOOLEAN DEFAULT 1,
    created_at DATETIME DEFAULT CURRENT_TIMESTAMP
);

-- =========================
-- FollowUp Templates Table
-- =========================
CREATE TABLE IF NOT EXISTS followup_templates (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    template_name VARCHAR(100),
    template_body TEXT
);

-- =========================
-- FollowUps Table
-- =========================
CREATE TABLE IF NOT EXISTS followups (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    client_id INTEGER,
    template_id INTEGER,
    scheduled_time DATETIME,
    created_by_employee VARCHAR(100),
    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
    FOREIGN KEY (client_id) REFERENCES clients(id),
    FOREIGN KEY (template_id) REFERENCES followup_templates(id)
);

-- =========================
-- Complaints Table
-- =========================
CREATE TABLE IF NOT EXISTS complaints (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    client_id INTEGER,
    message_text TEXT,
    is_resolved BOOLEAN DEFAULT 0,
    resolved_at DATETIME,
    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
    FOREIGN KEY (client_id) REFERENCES clients(id)
);

-- =========================
-- Motors Table
-- =========================
CREATE TABLE IF NOT EXISTS motors (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    english_name VARCHAR(40),
    arabic_name VARCHAR(40),
    company VARCHAR(40),
    agency_name VARCHAR(40),
    moto_type VARCHAR(40),
    price INTEGER,
    engin_capacity VARCHAR(40),
    fule_capacity VARCHAR(40),
    engin_type VARCHAR(40),
    transmission_type VARCHAR(40),
    max_speed VARCHAR(40),
    brake_type VARCHAR(40),
    colors VARCHAR(100),
    notes VARCHAR(100),
    is_available BOOLEAN DEFAULT 1,
    status VARCHAR(40),
    img_url VARCHAR(200)
);

-- =========================
-- Helmets Table
-- =========================
CREATE TABLE IF NOT EXISTS helmets (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    english_name VARCHAR(40),
    arabic_name VARCHAR(40),
    company VARCHAR(40),
    price INTEGER,
    helmet_type VARCHAR(40),
    colors VARCHAR(100),
    notes VARCHAR(100),
    is_available BOOLEAN DEFAULT 1,
    status VARCHAR(40),
    img_url VARCHAR(200)
);

-- =========================
-- Instalments Table
-- =========================
CREATE TABLE IF NOT EXISTS instalments (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    min_down_payment INTEGER,
    max_down_payment INTEGER,
    min_months INTEGER,
    max_months INTEGER,
    percentage FLOAT,
    percentage_per_month FLOAT
);

"""

# Instalment plans data (from the Ibyco dashboard)
INSTALMENT_INSERTS = [
    "INSERT INTO instalments (min_down_payment, max_down_payment, min_months, max_months, percentage, percentage_per_month) VALUES (50, 100, 0, 3, 0.0, 0.0);",
    "INSERT INTO instalments (min_down_payment, max_down_payment, min_months, max_months, percentage, percentage_per_month) VALUES (50, 100, 3, 6, 14.0, 2.3333);",
    "INSERT INTO instalments (min_down_payment, max_down_payment, min_months, max_months, percentage, percentage_per_month) VALUES (50, 100, 6, 12, 28.0, 2.3333);",
    "INSERT INTO instalments (min_down_payment, max_down_payment, min_months, max_months, percentage, percentage_per_month) VALUES (50, 100, 12, 18, 50.0, 2.7778);",
    "INSERT INTO instalments (min_down_payment, max_down_payment, min_months, max_months, percentage, percentage_per_month) VALUES (50, 100, 18, 24, 69.0, 2.875);",
    "INSERT INTO instalments (min_down_payment, max_down_payment, min_months, max_months, percentage, percentage_per_month) VALUES (30, 50, 0, 24, 72.0, 3.0);",
    "INSERT INTO instalments (min_down_payment, max_down_payment, min_months, max_months, percentage, percentage_per_month) VALUES (0, 30, 0, 24, 80.0, 3.3333);",
]


def esc(val):
    """Escape a value for SQL."""
    if val is None:
        return "NULL"
    s = str(val).replace("'", "''")
    return f"'{s}'"


def num(val):
    """Convert to integer for SQL."""
    if val is None:
        return "NULL"
    try:
        return str(int(float(val)))
    except (ValueError, TypeError):
        return "NULL"


def generate():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Sheet1"]

    motor_inserts = []
    helmet_inserts = []

    for row_idx in range(2, ws.max_row + 1):
        row = [cell.value for cell in ws[row_idx]]

        # Skip empty rows
        if not row[0] and not row[1]:
            continue

        # Excel column mapping:
        #  0: english_name    1: arabic_name     2: company
        #  3: agency_name     4: moto_type       5: price
        #  6: engin_capacity  7: engin_type      8: transmission_type
        #  9: max_speed      10: fule_capacity   11: brake_type
        # 12: notes          13-17: installment prices (ignored)
        # 18: img_url        19: colors          20: status
        # 21: is_available

        moto_type = row[4]
        img_url = row[18] if len(row) > 18 else None
        colors = row[19] if len(row) > 19 else None
        status = row[20] if len(row) > 20 else None

        # Helmets have moto_type = "خوزه"
        if moto_type and "خوز" in str(moto_type):
            # helmet_type is stored in engin_type col (نص / كامله)
            helmet_type = row[7]
            sql = (
                f"INSERT INTO helmets (english_name, arabic_name, company, price, "
                f"helmet_type, colors, notes, is_available, status, img_url) VALUES ("
                f"{esc(row[0])}, {esc(row[1])}, {esc(row[2])}, {num(row[5])}, "
                f"{esc(helmet_type)}, {esc(colors)}, {esc(row[12])}, 1, "
                f"{esc(status)}, {esc(img_url)});"
            )
            helmet_inserts.append(sql)
        else:
            sql = (
                f"INSERT INTO motors (english_name, arabic_name, company, agency_name, "
                f"moto_type, price, engin_capacity, fule_capacity, engin_type, "
                f"transmission_type, max_speed, brake_type, colors, notes, "
                f"is_available, status, img_url) VALUES ("
                f"{esc(row[0])}, {esc(row[1])}, {esc(row[2])}, {esc(row[3])}, "
                f"{esc(moto_type)}, {num(row[5])}, {esc(row[6])}, {esc(row[10])}, "
                f"{esc(row[7])}, {esc(row[8])}, {esc(row[9])}, {esc(row[11])}, "
                f"{esc(colors)}, {esc(row[12])}, 1, {esc(status)}, {esc(img_url)});"
            )
            motor_inserts.append(sql)

    # Write SQL file
    with open(SQL_PATH, "w", encoding="utf-8") as f:
        f.write(SCHEMA)

        f.write("-- =========================\n")
        f.write("-- Motors Data\n")
        f.write("-- =========================\n")
        for stmt in motor_inserts:
            f.write(stmt + "\n")

        f.write("\n-- =========================\n")
        f.write("-- Helmets Data\n")
        f.write("-- =========================\n")
        for stmt in helmet_inserts:
            f.write(stmt + "\n")

        f.write("\n-- =========================\n")
        f.write("-- Instalments Data\n")
        f.write("-- =========================\n")
        for stmt in INSTALMENT_INSERTS:
            f.write(stmt + "\n")

    print(f"Generated {SQL_PATH}")
    print(f"  Motors:       {len(motor_inserts)} rows")
    print(f"  Helmets:      {len(helmet_inserts)} rows")
    print(f"  Instalments:  {len(INSTALMENT_INSERTS)} rows")


if __name__ == "__main__":
    generate()
